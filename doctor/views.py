from pickle import FALSE
from typing import List
from xmlrpc.client import boolean
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import manager
from django.db.models.fields import NullBooleanField
from django.db.models.query import QuerySet
from django.forms.formsets import formset_factory
from django.shortcuts import render,redirect,get_object_or_404 
from django.http import HttpResponse ,Http404
import openpyxl
from doctor.models import BloodTest, Doctor, Patient
from datetime import datetime,timedelta,date
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
import logging    
from django.contrib import messages
import re
from openpyxl import load_workbook
import datetime as dt
import os
from django.core.files.storage import FileSystemStorage
from django.conf import settings
import simplejson as json
from django.shortcuts import render  
from django.http import HttpResponse 
import xlwt
from itertools import chain

def export_users_xls(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="medical_records.xls"'
 
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('גליון1')
 
    # Sheet header, first row
    row_num = 0
 
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['First Name', 'Last name', 'ID', 'Age', 'Gender','Smoking','EastCommunity','Pregnancy','Ethiopian','Diagnose','Treatment','WBC','Neut','Lymph','RBC','HCT','Urea','Hb','Crtn','Iron','HDL','AP']
 
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)
 
    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()
 
    patients = Patient.objects.all().values_list('name', 'lastName', 'patient_id', 'age','gender','smoke','eastCommunity','pregnancy','ethiopian','diagnose','treatment')
    blood_tests= BloodTest.objects.all().values_list('WBC','Neut','Lymph','RBC','HCT','Urea','Hb','Crtn','Iron','HDL','AP')
    rows=[patient + blood_test for patient, blood_test in zip(patients, blood_tests)]
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            if((col_num==12 or col_num==13 or col_num==15) and row[col_num]!=None):
                ws.write(row_num, col_num, row[col_num]+'%', font_style)
            else:
                ws.write(row_num, col_num, row[col_num], font_style)
    wb.save(response)
    return response
 

def get_homePage(request):
    return render(request,'index.html') #use html file
    
def homePageD(request,user_id):
    doctor = Doctor.objects.get(user_id = user_id)
    return render(request,'home.html',{'doctor':doctor}) #use html file
    

def logout_user(request):
    logout(request)
    return redirect('/')

#register to Doctor 
def submit_Doctor(request):
    user_id=request.POST.get('user_idup')
    name = request.POST.get('nameup')
    password = request.POST.get('passwordup')
    password2 = request.POST.get('password2up')

    if (CheckIfDoctorExist(user_id,name)):
        messages.error(request, 'קיים משתמש בעל אותה ת"ז או שם משתמש, אנא נסה שנית')
        return render(request,'index.html')

    else:
        if (not checkName(name)): 
            messages.error(request, 'ההרשמה לא בוצעה, שם משתמש לא תקין')
            return render(request,'index.html')
        if (len(user_id) != 9):
            messages.error(request, 'נא להזין ת"ז עם 9 ספרות (כולל ספרת ביקורת)')
            return render(request,'index.html')
        if (password != password2):
            messages.error(request, 'ההרשמה לא בוצעה, אנא הזן בשנית סיסמאות תואמות')
            return render(request,'index.html')
        if (not checkPassword(password)):
            print(password) 
            messages.error(request, 'ההרשמה לא בוצעה, סיסמה לא תקינה')
            return render(request,'index.html')

    doctor = Doctor(user_id = user_id,name=name,password=password)
    doctor.save()
    messages.success(request, "משתמש נוצר בהצלחה!")
    return render(request,'index.html')



def checkPassword(password):
    flag = False
    special = ['!','@','#','$','%','^','&','*','(',')','_','+','-','=','[',']','}',',',';',':','|','.','<','>','{','/','?','~', '\ ']
    if (len(password)<8 or len(password)>10):
        return False
    if (not ((re.search("[a-z]", password)) or (re.search("[A-Z]", password)))):
        return False
    if not re.search("[0-9]", password):
        return False
    for i in password:
        if i in special:
            flag = True
    return flag


def checkName(name):
    if (len(name)<6 or len(name)>8):
        return False
    count = 0
    for c in name:
        if c.isdigit():
            count += 1
        elif not c.isalpha():
            return False
    if count>2:
        return False
    return True



#Checks the username and the password for sign in
def validateDoctor(name,password): 
    if name!= None and password!= None:
        for i in Doctor.objects.all():
            if i.name == name and i.password == password:
                return True
    return False


def Conect(request): #conect the doctor to homepage
    name = request.POST.get('namein')
    password = request.POST.get('passwordin')

    if validateDoctor(name,password):
       doctor = Doctor.objects.get(name = name,password = password)
       return render(request,'home.html',{'doctor':doctor})
    else:
        messages.error(request, 'שם משתמש וסיסמה אינם נכונים, אנא נסה שנית')
        return render(request,'index.html')


def CheckIfDoctorExist(user_id,name):
    for i in Doctor.objects.all():
        if i.user_id == user_id :
            return True
        if i.name == name:
            return True
    return False
         

def patientQ(request,user_id):
    doctor = Doctor.objects.get(user_id = user_id)
    patient_id=request.POST.get('patient_id')
    gender = request.POST.get("gender")
    age = request.POST.get("age")
    name = request.POST.get("name")
    lastName = request.POST.get("lastName")
    smoke = request.POST.get("smoke")
    eastCommunity = request.POST.get("eastCommunity")
    ethiopian = request.POST.get("ethiopian")
    pregnancy = request.POST.get("pregnancy")


    if(patient_id!=None and gender!=None and age!=None and name!=None and lastName!=None):      
        #excel file
        bloodTest=BloodTest(WBC=None,Neut=None,Lymph=None,RBC=None,HCT=None,Urea=None,Hb=None,Crtn=None,Iron=None,HDL=None,AP=None)
        excel_file = request.FILES.get("excel_file",None)
        excel_data=None
        if (excel_file!=None):
            wb = openpyxl.load_workbook(excel_file)
            # getting all sheets
            sheets = wb.sheetnames

            # getting a particular sheet
            worksheet = wb["גיליון1"]

            # getting active sheet
            active_sheet = wb.active

            WBC = worksheet["A2"].value
            Neut = worksheet["B2"].value
            Lymph = worksheet["C2"].value  
            RBC = worksheet["D2"].value  
            HCT = worksheet["E2"].value  
            Urea = worksheet["F2"].value  
            Hb = worksheet["G2"].value  
            Crtn = worksheet["H2"].value         
            Iron = worksheet["I2"].value  
            HDL = worksheet["J2"].value  
            AP = worksheet["K2"].value  
            
            bloodTest=BloodTest(WBC=WBC,Neut=Neut,Lymph=Lymph,RBC=RBC,HCT=HCT,Urea=Urea,Hb=Hb,Crtn=Crtn,Iron=Iron,HDL=HDL,AP=AP)
            
        else:
            WBC = request.POST.get('WBC')
            Neut = request.POST.get('Neut')
            Lymph = request.POST.get('Lymph') 
            RBC = request.POST.get('RBC')
            HCT = request.POST.get('HCT')
            Urea = request.POST.get('Urea') 
            Hb = request.POST.get('Hb')
            Crtn = request.POST.get('Crtn')       
            Iron = request.POST.get('Iron')
            HDL = request.POST.get('HDL')
            AP = request.POST.get('AP')
            if(WBC=="" or Neut=="" or Lymph=="" or RBC=="" or HCT=="" or Urea=="" or
                Hb=="" or Crtn=="" or Iron=="" or HDL=="" or AP==""):
                    messages.error(request, 'אנא מלא את כל הערכים')
                    return render(request,'patientQ.html', {'doctor':doctor})

            else:
                bloodTest=BloodTest(WBC=WBC,Neut=Neut,Lymph=Lymph,RBC=RBC,HCT=HCT,Urea=Urea,Hb=Hb,Crtn=Crtn,Iron=Iron,HDL=HDL,AP=AP)
        bloodTest.save()
        
        lst=[smoke,eastCommunity,ethiopian,pregnancy]
        for i in range(0,len(lst)):
            if (lst[i]=="True"):
                lst[i]=True
            else:
                lst[i]=False
        patient = Patient(patient_id=patient_id, name=name, lastName= lastName,gender=gender,age=age,
        smoke=lst[0], pregnancy=lst[3], ethiopian=lst[2], eastCommunity=lst[1],bloodTest=bloodTest)
        

        if (age!= None):
            try:
                diagnose(bloodTest,patient,float(age))
            except:
                messages.error(request, 'אנא מלא את הפרטים תקין')
                return render(request,'patientQ.html', {'doctor':doctor})
            patient.save()

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="medical_record.xls"'
    
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('גליון1')
    
        # Sheet header, first row
        row_num = 0
        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columns = ['First Name', 'Last name', 'ID', 'Age', 'Gender','Smoking','EastCommunity','Pregnancy','Ethiopian','Diagnose','Treatment','WBC','Neut','Lymph','RBC','HCT','Urea','Hb','Crtn','Iron','HDL','AP']
    
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style)
    
        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()
        #excel for all the patient
        patients = Patient.objects.all().values_list('name', 'lastName', 'patient_id', 'age','gender','smoke','eastCommunity','pregnancy','ethiopian','diagnose','treatment')
        blood_tests= BloodTest.objects.all().values_list('WBC','Neut','Lymph','RBC','HCT','Urea','Hb','Crtn','Iron','HDL','AP')
        rows=[patient1 + blood_test for patient1, blood_test in zip(patients, blood_tests)]
        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                if (row_num == len(rows)):
                    if((col_num==12 or col_num==13 or col_num==15) and row[col_num]!=None):
                        ws.write(1, col_num, row[col_num]+'%', font_style)
                    elif(col_num==9):
                        diagnoselst=patient.diagnose.split('""')
                        for i in range(1,len(diagnoselst),1):
                            ws.write(i, 9, diagnoselst[i], font_style)
                    elif(col_num==10):   
                        treatlst=patient.treatment.split('""')
                        for i in range(1,len(treatlst),1):
                            ws.write(i, 10, treatlst[i], font_style) 
                    else:
                        ws.write(1, col_num, row[col_num], font_style)
        wb.save(response)
        return response
        #return render(request,'patientQ.html', {'doctor':doctor,'patient':patient})
    else:
        return render(request,'patientQ.html', {'doctor':doctor})


def addPatientSucc(request,user_id):
    doctor = Doctor.objects.get(user_id = user_id)
    return render(request,'sucsees.html',{'doctor' :doctor})

def checkTrack(track,lst):
    for i in lst:
        if i not in track:
            track.append(i)

def diagnose(bloodtest,patient,age):
    anemia  = " אנמיה- שני כדורי 10 מג של בי12 ביום למשך חודש "
    diet = " דיאטה- לתאם פגישה עם תזונאי "
    bleading = " דימום- להתפנות בדחיפות לבית החולים "
    hiperlipidemia = " היפרליפידמיה(שומנים בדם)- של לתאם פגישה עם תזונאי, כדור 5 מג של סימוביל ביום למשך שבוע "
    disruptionOfBlood = " הפרעה ביצירת הדם\תאי הדם - כדור 10 מג של בי12 ביום למשך חודש וכדור 5 מג של חומצה פולית ביום למשך חודש "
    hematologicalDisorder = " הפרעה המטולוגיה- זריקה של הורמון לעידוד ייצור תאי הדם האדומים "
    ironPoisoning= " הרעלת ברזל - להתפנות לבית החולים "
    dehydration = " התייבשות- מנוחה מוחלטת בשכיבה, החזרת נוזלים בשתייה "
    infection = " זיהום- אנטיביוטיקה ייעודית "
    vitaminDef = " חוסר בוויטמינם- הפנייה לבדיקת דם לזיהוי הוויטמינים החסרים "
    viralDisease = " מחלה ויראלית- לנוח בבית "
    diseaseInBile = " מחלות בדרכי המרה- הפנייה לטיפול כירורגי "
    heartDisease = " מחלות לב - לתאם פגישה עם תזונאי "
    bloodDisease = " מחלת דם- שילוב של ציקלופוספאמיד וקורטיקוסרואידים "
    liverDisease = " מחלת כבד- הפנייה לאבחנה ספציפית לצורך קביעת טיפול "
    kidneyDisease = " מחלת כליה - איזון את רמות הסוכר בדם "
    ironDef = " מחסור בברזל - שני כדורי 10 מג של בי12 ביום למשך חודש "
    muscleDisease = " מחלות שריר- שני כדורי 5 מג של כורכום סי3 של אלטמן ביום למשך חודש "
    smokeing = " מעשנים- להפסיק לעשן "
    lungDisease = " מחלת ריאות- להפסיק לעשן / הפנייה לצילום רנטגן של הריאות "
    overactiveThyroid = "  פעילות יתר של בלוטת התריס- יש לקחת Propylthiouracil להקטנת פעילות בלוטת התריס "
    adultDiabetes = " סוכרת מבוגרים- התאמת אינסולין למטופל "
    cancer = " אנטרקטיניב - סרטן "
    meatInc = " צריכה מוגברת של בשר- לתאם פגישה עם תזונאי "
    variousMedications = " שימוש בתרופות שונות- הפנייה לרופא המשפחה לצורך בדיקת התאמה בין התרופות "
    malnutrition = " תת תזונה- לתאם פגישה עם תזונאי "
    patient.diagnose=json.dumps(" ")
    patient.treatment=json.dumps(" ")
    track = []
    #WBC check
    if (age!=None):
        if (age>=0 and age<=3):
            if (float(bloodtest.WBC)<6000):
                patient.diagnose+=json.dumps(" ערכים נמוכים של כמות תאי הדם הלבנים הכללית מצביעים על מחלה ויראלית, כשל של מערכת החיסון ובמקרים נדירים ביותר על סרטן. ",ensure_ascii=False)
                checkTrack(track, [viralDisease,cancer,bloodDisease])
            if (float(bloodtest.WBC)>17500):
                patient.diagnose+=json.dumps(" ערכים גבוהים של כמות תאי הדם הלבנים הכללית מצביעים לרוב על קיומו של זיהום, אם קיימת מחלת חום. במקרים אחרים, נדירים ביותר, עלולים ערכים גבוהים מאוד להעיד על מחלת דם או סרטן. ",ensure_ascii=False)
                checkTrack(track, [infection,bloodDisease,cancer])
        elif (age>3 and age<=17):
            if (float(bloodtest.WBC)<5500):
                patient.diagnose+=json.dumps(" ערכים נמוכים של כמות תאי הדם הלבנים הכללית מצביעים על מחלה ויראלית, כשל של מערכת החיסון ובמקרים נדירים ביותר על סרטן. ",ensure_ascii=False)
                checkTrack(track, [viralDisease,bloodDisease,cancer])           
            if (float(bloodtest.WBC)>15500):
                patient.diagnose+=json.dumps(" ערכים גבוהים של כמות תאי הדם הלבנים הכללית מצביעים לרוב על קיומו של זיהום, אם קיימת מחלת חום. במקרים אחרים, נדירים ביותר, עלולים ערכים גבוהים מאוד להעיד על מחלת דם או סרטן. ",ensure_ascii=False)
                checkTrack(track, [infection,bloodDisease,cancer])
        else:
            if (float(bloodtest.WBC)<4500):
                patient.diagnose+=json.dumps(" ערכים נמוכים של כמות תאי הדם הלבנים הכללית מצביעים על מחלה ויראלית, כשל של מערכת החיסון ובמקרים נדירים ביותר על סרטן. ",ensure_ascii=False)
                checkTrack(track, [viralDisease,bloodDisease,cancer])       
            if (float(bloodtest.WBC)>11000):
                patient.diagnose+=json.dumps(" ערכים גבוהים של כמות תאי הדם הלבנים הכללית מצביעים לרוב על קיומו של זיהום, אם קיימת מחלת חום. במקרים אחרים, נדירים ביותר, עלולים ערכים גבוהים מאוד להעיד על מחלת דם או סרטן. ",ensure_ascii=False)
                checkTrack(track, [infection,bloodDisease,cancer])       
    
    #Neut check
    if(float(bloodtest.Neut)<28):
        patient.diagnose+=json.dumps(" ערכים נמוכים בכמות תאי הדם הלבנים האחראים בעיקר על חיסול החיידקים- מעידים על הפרעה ביצירת הדם, על נטייה לזיהומים מחיידקים ובמקרים נדירים - על תהליך סרטני. ",ensure_ascii=False)
        checkTrack(track, [disruptionOfBlood,infection,cancer])            
    if(float(bloodtest.Neut)>54):
        patient.diagnose+=json.dumps(" ערכים גבוהים בכמות תאי הדם הלבנים האחראים בעיקר על חיסול החיידקים מעידים על זיהום חיידקי. ",ensure_ascii=False)
        checkTrack(track, [infection])                          
    
    #Lymph check
    if(float(bloodtest.Lymph)<36):
        patient.diagnose+=json.dumps(" ערכים נמוכים תאי דם לבנים האחראים על הריגת נגיפים או חיידקים הנמצאים בגוף זמן ממושך מעידים על בעיה ביצירת תאי דם ",ensure_ascii=False)
        checkTrack(track, [disruptionOfBlood]) 
    if(float(bloodtest.Lymph)>52):    
        patient.diagnose+=json.dumps(" ערכים גבוהים בתאי דם לבנים האחראים על הריגת נגיפים או חיידקים הנמצאים בגוף זמן ממושך עשויים להצביע על זיהום חידקי ממושך או על סרטן הלימפומה ",ensure_ascii=False)        
        checkTrack(track, [infection,cancer]) 
    
    #RBC check
    if(float(bloodtest.RBC)<4.5):
        patient.diagnose+=json.dumps(" ערכים נמוכים בכדוריות הדם האדומות אחראיות על קשירת חמצן מהריאות, על הובלתו לרקמות הגוף, על קליטת פחמן דו-חמצנימתאי הגוף השונים ועל פליטתו בחזרה לריאות - עלולים להצביע על אנמיה או על דימומים קשים. ",ensure_ascii=False)
        checkTrack(track, [anemia,bleading]) 
    if(float(bloodtest.RBC)>6):    
        patient.diagnose+=json.dumps(" ערכים נמוכים בכדוריות הדם האדומות אחראיות על קשירת חמצן מהריאות, על הובלתו לרקמות הגוף, על קליטת פחמן דו-חמצנימתאי הגוף השונים ועל פליטתו בחזרה לריאות -  עלולים להצביע על הפרעה במערכת ייצור הדם. רמות גבוהות נצפו גם אצל מעשנים ואצל חולים במחלות ריאות. ",ensure_ascii=False)
        checkTrack(track, [lungDisease,disruptionOfBlood]) 
        if (patient.smoke == True):
            checkTrack(track, [smokeing]) 

    #HCT check
    if(patient.gender == "woman"):
        if (float(bloodtest.HCT)<33):
            patient.diagnose+=json.dumps(" ערכים נמוכים בנפח כדוריות הדם האדומות בתוך כלל נוזל הדם. - מצביעים לרוב על דימום או על אנמיה ",ensure_ascii=False)
            checkTrack(track, [anemia,bleading]) 
        if (float(bloodtest.HCT)>47):
            patient.diagnose+=json.dumps(" ערכים נמוכים בנפח כדוריות הדם האדומות בתוך כלל נוזל הדם. - שכיח בדרך כלל אצל מעשנים. ",ensure_ascii=False)
        if (patient.smoke == True):
            checkTrack(track, [smokeing]) 
    if(patient.gender == "man"):
        if (float(bloodtest.HCT)<37):
            patient.diagnose+=json.dumps(" ערכים נמוכים בנפח כדוריות הדם האדומות בתוך כלל נוזל הדם. - מצביעים לרוב על דימום או על אנמיה ",ensure_ascii=False)
            checkTrack(track, [anemia,bleading]) 
        if (float(bloodtest.HCT)>54):
            patient.diagnose+=json.dumps(" ערכים נמוכים בנפח כדוריות הדם האדומות בתוך כלל נוזל הדם. - שכיח בדרך כלל אצל מעשנים.  ",ensure_ascii=False)
        if (patient.smoke == True):
            checkTrack(track, [smokeing]) 

    #Urea check
    if((patient.eastCommunity == False and float(bloodtest.Urea)<17) or (patient.eastCommunity == True and float(bloodtest.Urea)<18.7)):
        patient.diagnose+=json.dumps(" ערכים נמוכים ברמת השתנן בדם. שתנן הוא התוצר הסופי של תהליך חילוף החומרים של חלבונים בגוף- עלול להצביע על: תת תזונה, דיאטה דלת חלבון או מחלת כבד. ",ensure_ascii=False)
        if(patient.pregancy == True):
            patient.diagnose+=json.dumps(" המטופלת בהריון - נא לשים לב בהריון רמת השתנן יורדת. ",ensure_ascii=False)
        checkTrack(track, [malnutrition,diet,liverDisease]) 
    if((patient.eastCommunity == False and float(bloodtest.Urea)>43) or (patient.eastCommunity == True and float(bloodtest.Urea)>47.3)):
        patient.diagnose+=json.dumps(" ערכים גבוהים ברמת השתנן בדם. שתנן הוא התוצר הסופי של תהליך חילוף החומרים של חלבונים בגוף - עלולים להצביע על מחלות כליה,התייבשות או דיאטה עתירת חלבונים ",ensure_ascii=False)
        checkTrack(track, [kidneyDisease,dehydration,diet]) 

    #HB check
    if((patient.gender == "woman" and age >17 and float(bloodtest.Hb)<12) or (patient.gender == "man" and age >17 and float(bloodtest.Hb)<12) or (age <=17 and float(bloodtest.Hb)<11.5)):
        patient.diagnose+=json.dumps(" ערכים נמוכים ברמת ההמוגלובין- המוגלובין הוא מרכיב בתוך הכדורית האדומה, אשר אחראי על קשירתם ועל שחרורם של חמצן ושל פחמן דו-חמצני-מעידים על אנמיה. זו יכולה לנבוע מהפרעה המטולוגית, ממחסור בברזל ומדימומים. ",ensure_ascii=False)
        checkTrack(track, [anemia,hematologicalDisorder,ironDef,bleading]) 

    if((patient.gender == "woman" and age >17 and float(bloodtest.Hb)>16) or (patient.gender == "man" and age >17 and float(bloodtest.Hb)>18) or (age <=17 and float(bloodtest.Hb)>15.5)):
        patient.diagnose+=json.dumps(" ערכים גבוהים ברמת ההמוגלובין- המוגלובין הוא מרכיב בתוך הכדורית האדומה . ",ensure_ascii=False)

    #Crtn check
    if((age>=0 and age<=2 and float(bloodtest.Crtn)<0.2) or (age>=3 and age<=17 and float(bloodtest.Crtn)<0.5) or (age>=18 and age<=59 and float(bloodtest.Crtn)<0.6) or (age>=60 and float(bloodtest.Crtn)<0.6)):
        patient.diagnose+=json.dumps(" ערכים נמוכים בקריטאינין -  תוצר פירוק של מרכיב המיוצר בגוף ונמצא בשריר הקרוי קריאנין פוספט. בדיקת קריאטינין חשובה ביותר כיוון שהיא נותנת אמת מידה לגבי תפקוד הכליות. ערכים נמוכים נראים לרוב בחולים בעלי מסת שריר ירודה מאוד ואנשים בתת תזונה שאינם צורכים די חלבון",ensure_ascii=False)
        checkTrack(track, [anemia,muscleDisease,malnutrition]) 
    if((age>=0 and age<=2 and float(bloodtest.Crtn)>0.5) or (age>=3 and age<=17 and float(bloodtest.Crtn)>1) or (age>=18 and age<=59 and float(bloodtest.Crtn)>1) or (age>=60 and float(bloodtest.Crtn)>1.2)):
        patient.diagnose+=json.dumps(" ערכים גבוהים בקריטאינין - תוצר פירוק של מרכיב המיוצר בגוף ונמצא בשריר הקרוי קריאנין פוספט. בדיקת קריאטינין חשובה ביותר כיוון שהיא נותנת אמת מידה לגבי תפקוד הכליות- הערכים עלולים להצביע על בעיה כלייתית ובמקרים חמורים על אי ספיקת כליות. ערכים גבוהים ניתן למצוא גם בעת שלשולים והקאות (הגורמים לפירוק מוגבר של שריר ולערכים גבוהים של קריאטינין), מחלות שריר וצריכה מוגברת של בשר. ",ensure_ascii=False)
        checkTrack(track, [kidneyDisease,muscleDisease,meatInc]) 


    #Iron check
    if ((patient.gender == "man" and float(bloodtest.Iron)<60) or (patient.gender == "woman" and float(bloodtest.Iron)<48)):
        patient.diagnose+=json.dumps(" ערכים נמוכים בברזל - הברזל חיוני ליצירת ההמוגלובין - החלבון שנושא את החמצן בדם. נוסף על כך הוא משמש ליצירת אנזימים רבים אחרים. רמות נמוכות של ברזל מעידה בדרך כלל על תזונה לא מספקת או על עלייה בצורך בברזל (למשל בהריון) או על איבוד דם בעקבות דימום. ",ensure_ascii=False)
        checkTrack(track, [diet,bleading,ironDef]) 
    if ((patient.gender == "man" and float(bloodtest.Iron)>160) or (patient.gender == "woman" and float(bloodtest.Iron)>128)):
        patient.diagnose+=json.dumps(" ערכים גבוהים בברזל - הברזל חיוני ליצירת ההמוגלובין - החלבון שנושא את החמצן בדם. נוסף על כך הוא משמש ליצירת אנזימים רבים אחרים. רמות גבוהות של עלולים להצביע על הרעלת ברזל ",ensure_ascii=False)
        checkTrack(track, [ironPoisoning])

    #HDL check
    if ((patient.ethiopian == False and (patient.gender == "man" and float(bloodtest.HDL)<29) or (patient.gender == "woman" and float(bloodtest.HDL)<34)) or (patient.ethiopian == True and (patient.gender == "man" and float(bloodtest.HDL)<34.8) or (patient.gender == "woman" and float(bloodtest.HDL)<40.8)) ):
        patient.diagnose+=json.dumps(" ערכים נמוכים באץ'-די-אל הקרוי גם הכולסטרול הטוב, הינו מולקולה דמוית חלבון, אשר נושאת את הכולסטרול מתאי הגוף אל הכבד,שם מפורק הכולסטרול. בכך מסייע ה-אץ'-די-אל לגוף להיפטר מעודפי שומנים. רמות נמוכותת עשויות להצביע על סיכון למחלות לב, על היפרליפידמיה (יתר שומנים בדם) או על סוכרת מבוגרים. ",ensure_ascii=False)
        checkTrack(track, [heartDisease,hiperlipidemia,adultDiabetes]) 
    if ((patient.ethiopian == False and (patient.gender == "man" and float(bloodtest.HDL)>62) or (patient.gender == "woman" and float(bloodtest.HDL)>82)) or (patient.ethiopian == True and (patient.gender == "man" and float(bloodtest.HDL)>62*1.2) or (patient.gender == "woman" and float(bloodtest.HDL)>82*1.2)) ):
        patient.diagnose+=json.dumps(" ערכים גבוהים באץ'-די-אל הקרוי גם הכולסטרול הטוב, הינו מולקולה דמוית חלבון, אשר נושאת את הכולסטרול מתאי הגוף אל הכבד,שם מפורק הכולסטרול. בכך מסייע ה-אץ'-די-אל לגוף להיפטר מעודפי שומנים. רמות גבוהות לרוב אינן מזיקות. פעילות גופנית מעלה את רמות הכולסטרול הטוב. ",ensure_ascii=False)

    #AP check
    if ((patient.eastCommunity == False and  float(bloodtest.AP)<30) or (patient.eastCommunity == True and  float(bloodtest.AP)<60)):
        patient.diagnose+=json.dumps(" ערכים נמוכים בפוספטזה אלקלית - התפקיד המטבולי של האנזים אינו ברור לגמרי. הוא קשור להעברת מרכיבים דרך ממברנות, וכן יש לו תפקיד בהסתיידות של העצם. האנזים נמצא ברקמות שונות בגוף, בעיקר בכבד, בדרכי המרה, במעי, בעצמות ובשליה. רמות נמוכות עשויות להצביע על תזונה לקויה שחסרים בה חלבונים. חוסר בוויטמינים כמו ויטמין בי12 , ויטמין סי, ויטמין בי6 וחומצה פולית. ",ensure_ascii=False)
        checkTrack(track, [diet,vitaminDef]) 
    if ((patient.eastCommunity == False and  float(bloodtest.AP)>90) or (patient.eastCommunity == True and  float(bloodtest.AP)>120)):
        patient.diagnose+=json.dumps(" ערכים גבוהים  בפוספטזה אלקלית - התפקיד המטבולי של האנזים אינו ברור לגמרי. הוא קשור להעברת מרכיבים דרך ממברנות, וכן יש לו תפקיד בהסתיידות של העצם. האנזים נמצא ברקמות שונות בגוף, בעיקר בכבד, בדרכי המרה, במעי, בעצמות ובשליה. רמות גבוהות עלולים להצביע על מחלות כבד, מחלות בדרכי המרה, הריון, פעילות יתר של בלוטת התריס או שימוש בתרופות שונות ",ensure_ascii=False)
        checkTrack(track, [liverDisease,diseaseInBile,overactiveThyroid,variousMedications]) 
        if (patient.pregnancy == True):
            patient.treatment+=json.dumps(" המטופלת בהריון - נא לשים לב יכולות להיות רמות גבוהות בגלל ההריון ",ensure_ascii=False)
        if (patient.pregnancy == False and patient.gender == "woman"):
            patient.treatment+=json.dumps(" עלול להיגרם בגלל הריון - הפנייה לשלילת הריון ",ensure_ascii=False)
    for i in track:
        patient.treatment+=json.dumps(i,ensure_ascii=False)

